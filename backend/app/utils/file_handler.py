import io
from pathlib import Path
from typing import Union, Optional, List, Dict, Any
import pandas as pd
import chardet
import structlog
import openpyxl
from openpyxl.utils import get_column_letter

logger = structlog.get_logger()


class FileHandler:
    def __init__(self):
        self.encodings_to_try = ["utf-8", "latin-1", "cp1252", "iso-8859-1"]
        self._workbook = None
        self._current_file_path = None
    
    async def read_file(self, file_path: Union[str, Path]) -> pd.DataFrame:
        file_path = Path(file_path)
        
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        file_extension = file_path.suffix.lower()
        
        if file_extension == ".csv":
            return await self._read_csv(file_path)
        elif file_extension in [".xlsx", ".xls"]:
            return await self._read_excel(file_path)
        else:
            raise ValueError(f"Unsupported file type: {file_extension}")
    
    async def read_file_contents(
        self,
        contents: bytes,
        filename: str
    ) -> pd.DataFrame:
        file_extension = Path(filename).suffix.lower()
        
        if file_extension == ".csv":
            return await self._read_csv_contents(contents)
        elif file_extension in [".xlsx", ".xls"]:
            return await self._read_excel_contents(contents)
        else:
            raise ValueError(f"Unsupported file type: {file_extension}")
    
    async def _read_csv(self, file_path: Path) -> pd.DataFrame:
        # Detect encoding
        encoding = self._detect_encoding(file_path)
        
        # Try to read with detected encoding
        for enc in [encoding] + self.encodings_to_try:
            try:
                df = pd.read_csv(file_path, encoding=enc)
                logger.info(f"Successfully read CSV with encoding: {enc}")
                return self._clean_dataframe(df)
            except UnicodeDecodeError:
                continue
            except Exception as e:
                logger.error(f"Error reading CSV: {e}")
                raise
        
        raise ValueError("Could not decode CSV file with any supported encoding")
    
    async def _read_csv_contents(self, contents: bytes) -> pd.DataFrame:
        # Detect encoding from contents
        detection = chardet.detect(contents)
        encoding = detection.get("encoding", "utf-8")
        
        for enc in [encoding] + self.encodings_to_try:
            try:
                df = pd.read_csv(io.BytesIO(contents), encoding=enc)
                return self._clean_dataframe(df)
            except UnicodeDecodeError:
                continue
            except Exception as e:
                logger.error(f"Error reading CSV contents: {e}")
                raise
        
        raise ValueError("Could not decode CSV contents with any supported encoding")
    
    async def _read_excel(self, file_path: Path, sheet_name: Union[str, int, None] = None) -> pd.DataFrame:
        try:
            # Store the file path for sheet operations
            self._current_file_path = file_path
            
            # If no sheet specified, read the first sheet
            if sheet_name is None:
                sheet_name = 0
            
            df = pd.read_excel(file_path, sheet_name=sheet_name, engine="openpyxl")
            return self._clean_dataframe(df)
        except Exception as e:
            logger.error(f"Error reading Excel file: {e}")
            raise
    
    async def _read_excel_contents(self, contents: bytes, sheet_name: Union[str, int, None] = None) -> pd.DataFrame:
        try:
            # If no sheet specified, read the first sheet
            if sheet_name is None:
                sheet_name = 0
                
            df = pd.read_excel(io.BytesIO(contents), sheet_name=sheet_name, engine="openpyxl")
            return self._clean_dataframe(df)
        except Exception as e:
            logger.error(f"Error reading Excel contents: {e}")
            raise
    
    def _detect_encoding(self, file_path: Path) -> str:
        try:
            with open(file_path, "rb") as f:
                raw_data = f.read(10000)  # Read first 10KB
                result = chardet.detect(raw_data)
                return result.get("encoding", "utf-8")
        except Exception:
            return "utf-8"
    
    def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        # Remove completely empty rows and columns
        df = df.dropna(how="all", axis=0)
        df = df.dropna(how="all", axis=1)
        
        # Strip column names
        df.columns = [str(col).strip() for col in df.columns]
        
        # Remove unnamed columns (often index columns from Excel)
        unnamed_cols = [col for col in df.columns if col.startswith("Unnamed:")]
        if unnamed_cols:
            df = df.drop(columns=unnamed_cols)
        
        return df
    
    async def save_dataframe(
        self,
        df: pd.DataFrame,
        file_path: Union[str, Path],
        format: str = "csv"
    ):
        file_path = Path(file_path)
        
        if format == "csv":
            df.to_csv(file_path, index=False, encoding="utf-8")
        elif format == "xlsx":
            df.to_excel(file_path, index=False, engine="openpyxl")
        elif format == "json":
            df.to_json(file_path, orient="records", indent=2)
        else:
            raise ValueError(f"Unsupported format: {format}")
        
        logger.info(f"Saved dataframe to {file_path} as {format}")
    
    async def get_excel_sheets(self, file_path: Union[str, Path, bytes]) -> List[Dict[str, Any]]:
        """Get information about all sheets in an Excel file"""
        sheets_info = []
        
        try:
            if isinstance(file_path, bytes):
                workbook = openpyxl.load_workbook(io.BytesIO(file_path), read_only=True, data_only=True)
            else:
                file_path = Path(file_path)
                workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
            for idx, sheet_name in enumerate(workbook.sheetnames):
                sheet = workbook[sheet_name]
                
                # Get sheet dimensions
                max_row = sheet.max_row
                max_col = sheet.max_column
                
                # Count non-empty rows
                non_empty_rows = 0
                for row in sheet.iter_rows(max_row=min(100, max_row)):
                    if any(cell.value is not None for cell in row):
                        non_empty_rows += 1
                
                sheets_info.append({
                    "index": idx,
                    "name": sheet_name,
                    "rows": max_row,
                    "columns": max_col,
                    "non_empty_rows": non_empty_rows,
                    "hidden": sheet.sheet_state == "hidden"
                })
            
            workbook.close()
            return sheets_info
            
        except Exception as e:
            logger.error(f"Error getting Excel sheets: {e}")
            return []
    
    async def preview_excel_sheet(self, file_path: Union[str, Path, bytes], sheet_name: Union[str, int] = 0, rows: int = 20) -> Dict[str, Any]:
        """Preview data from a specific Excel sheet"""
        try:
            if isinstance(file_path, bytes):
                df = pd.read_excel(io.BytesIO(file_path), sheet_name=sheet_name, engine="openpyxl", nrows=rows)
            else:
                df = pd.read_excel(file_path, sheet_name=sheet_name, engine="openpyxl", nrows=rows)
            
            # Clean the dataframe
            df = self._clean_dataframe(df)
            
            # Get column info
            columns_info = []
            for col in df.columns:
                col_data = df[col].dropna()
                columns_info.append({
                    "name": col,
                    "type": str(df[col].dtype),
                    "non_null_count": len(col_data),
                    "unique_count": col_data.nunique() if len(col_data) > 0 else 0,
                    "sample_values": col_data.head(5).tolist() if len(col_data) > 0 else []
                })
            
            return {
                "headers": df.columns.tolist(),
                "data": df.to_dict(orient="records"),
                "columns_info": columns_info,
                "total_rows": len(df),
                "sheet_name": sheet_name if isinstance(sheet_name, str) else f"Sheet{sheet_name + 1}"
            }
            
        except Exception as e:
            logger.error(f"Error previewing Excel sheet: {e}")
            raise
    
    async def read_excel_with_structure(self, file_path: Union[str, Path, bytes], sheet_name: Union[str, int] = 0) -> Dict[str, Any]:
        """Read Excel file and preserve structure information like merged cells, formatting, etc."""
        try:
            if isinstance(file_path, bytes):
                workbook = openpyxl.load_workbook(io.BytesIO(file_path), data_only=True)
            else:
                workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            # Get the sheet
            if isinstance(sheet_name, int):
                sheet = workbook.worksheets[sheet_name]
            else:
                sheet = workbook[sheet_name]
            
            # Extract data with structure
            data = []
            headers = []
            merged_cells = []
            
            # Get merged cell ranges
            for merged_range in sheet.merged_cells.ranges:
                merged_cells.append({
                    "range": str(merged_range),
                    "top_left": {
                        "row": merged_range.min_row,
                        "col": merged_range.min_col,
                        "value": sheet.cell(merged_range.min_row, merged_range.min_col).value
                    }
                })
            
            # Read all data
            all_data = []
            for row in sheet.iter_rows():
                row_data = []
                for cell in row:
                    cell_info = {
                        "value": cell.value,
                        "row": cell.row,
                        "column": cell.column,
                        "column_letter": cell.column_letter,
                        "data_type": cell.data_type,
                        "is_merged": any(cell.coordinate in merged_range for merged_range in sheet.merged_cells.ranges)
                    }
                    row_data.append(cell_info)
                all_data.append(row_data)
            
            workbook.close()
            
            # Try to identify headers (usually first non-empty row)
            header_row_idx = 0
            for idx, row in enumerate(all_data):
                if any(cell["value"] is not None for cell in row):
                    header_row_idx = idx
                    headers = [cell["value"] if cell["value"] is not None else f"Column_{i+1}" 
                              for i, cell in enumerate(row)]
                    break
            
            # Convert to structured format
            structured_data = {
                "raw_data": all_data,
                "headers": headers,
                "header_row_index": header_row_idx,
                "merged_cells": merged_cells,
                "sheet_name": sheet.title,
                "dimensions": {
                    "rows": sheet.max_row,
                    "columns": sheet.max_column
                }
            }
            
            return structured_data
            
        except Exception as e:
            logger.error(f"Error reading Excel with structure: {e}")
            raise